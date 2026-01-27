import logging
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from rest_framework.renderers import BaseRenderer, TemplateHTMLRenderer

logger = logging.getLogger(__name__)


class PrintRenderer(TemplateHTMLRenderer):
    format = "print"

    def get_template_names(self, response, view):
        view_name = view.__class__.__name__.lower()
        template_name = f"print-templates/{view_name}.html"
        templates = [template_name, "print-templates/_default.html"]
        logger.info(templates)
        return templates


class XLSXRenderer(BaseRenderer):
    """
    Renderer for Excel spreadsheet open data format (xlsx).
    """

    media_type = "application/xlsx"
    format = "xlsx"
    exclude_fields = ("id",)
    digits_fields = (
        "plan",
        "corrected_plan",
        "contract_total",
        "remains",
        "pay_total",
        "total",
    )

    def render(self, data, accepted_media_type=None, renderer_context=None):
        row_count = 0

        if data is None:
            return bytes()

        wb = Workbook()

        ws = wb.active
        ws.title = "Export"

        for row in data["results"]:
            column_count = 0
            row_count += 1

            for key, value in row.items():
                if key in self.exclude_fields:
                    continue
                column_count += 1
                if key in self.digits_fields:
                    value = Decimal(value.replace(" ", ""))
                else:
                    value = str(value or "")

                ws.cell(
                    row=row_count,
                    column=column_count,
                    value=value,
                )

        for ws_column in range(1, column_count + 1):
            col_letter = get_column_letter(ws_column)
            ws.column_dimensions[col_letter].width = 15

        return save_virtual_workbook(wb)
